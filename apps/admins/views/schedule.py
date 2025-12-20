from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication


from rest_framework.parsers import MultiPartParser, FormParser
import openpyxl
from dataclasses import dataclass
import logging
from datetime import date, datetime, timedelta

from apps.my_built_in.response import ResponseFormat
from apps.admins.serializers.schedule import (
    ScheduleInputSerializer,
    ScheduleResultSerializer,
    CourseScheduleSerializer,
    HardAssignmentSerializer
)
from apps.admins.services.scheduler import GeneticScheduler
from apps.my_built_in.models.lop_tin_chi import LopTinChi
from django.db import transaction

logger = logging.getLogger(__name__)


@dataclass
class HardAssignment:
    """Phân công cứng - Bắt BUỘC phải tuân theo"""
    id: int
    course_id: int
    teacher_id: int
    is_active: bool = True
    is_deleted: bool = False


class ScheduleView(APIView):
    """
    API xếp lịch học tự động bằng Genetic Algorithm
    """
    parser_classes = (MultiPartParser, FormParser)
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Chạy thuật toán xếp lịch

        Request body:
        - semester_id: ID học kỳ cần xếp lịch
        - excel_file: File Excel chứa hard assignments (optional)
        - population_size: Kích thước quần thể (optional, default=100)
        - generations: Số thế hệ tối đa (optional, default=200)
        - crossover_rate: Tỷ lệ lai ghép (optional, default=0.8)
        - mutation_rate: Tỷ lệ đột biến (optional, default=0.05)
        """

        # Validate input
        serializer = ScheduleInputSerializer(data=request.data)
        if not serializer.is_valid():
            return ResponseFormat.response(
                data=serializer.errors,
                case_name="INVALID_INPUT",
                status=400
            )

        validated_data = serializer.validated_data
        semester_id = validated_data['semester_id']
        excel_file = validated_data.get('excel_file')
        holiday_file = validated_data.get('holiday_file', [])

        try:
            # Parse hard assignments từ Excel nếu có
            hard_assignments = []
            if excel_file:
                logger.info("Đang đọc file Excel hard assignments...")
                hard_assignments = self._parse_excel_hard_assignments(excel_file)
                if holiday_file:
                    holidays = self._parse_excel_holidays(holiday_file)
                else:
                    holidays = []
                logger.info(f"Đã đọc {len(hard_assignments)} phân công cứng từ Excel")
            # print("hard_assignments",hard_assignments)
            # Khởi tạo scheduler với tham số tùy chỉnh
            scheduler = GeneticScheduler(semester_id=semester_id, HOLIDAYS=holidays)

            # Override parameters nếu có
            if 'population_size' in validated_data:
                scheduler.POPULATION_SIZE = validated_data['population_size']
            if 'generations' in validated_data:
                scheduler.GENERATIONS = validated_data['generations']
            if 'crossover_rate' in validated_data:
                scheduler.CROSSOVER_RATE = validated_data['crossover_rate']
            if 'mutation_rate' in validated_data:
                scheduler.MUTATION_RATE = validated_data['mutation_rate']

            # Lấy danh sách lớp cần xếp
            courses = scheduler.get_courses_to_schedule()

            if not courses:
                return ResponseFormat.response(
                    data={
                        'success': True,
                        'message': 'Không có lớp nào cần xếp lịch',
                        'scheduled_courses': [],
                        'statistics': {},
                        'violations': {}
                    },
                    case_name="SUCCESS"
                )

            # Khởi tạo quần thể
            scheduler.initialize_population(courses, hard_assignments)

            # Chạy thuật toán
            logger.info(f"Bắt đầu chạy GA với {len(courses)} lớp và {len(hard_assignments)} phân công cứng")
            result = scheduler.run(hard_assignments)

            if not result['success'] or not result['best_chromosome']:
                return ResponseFormat.response(
                    data={'message': 'Không thể tạo lịch học hợp lệ'},
                    case_name="ERROR",
                    status=500
                )

            best_chromosome = result['best_chromosome']

            # Áp dụng lịch học vào database
            logger.info("Đang lưu lịch học vào database...")
            saved = scheduler.save_to_database(best_chromosome)

            if not saved:
                return ResponseFormat.response(
                    data={'message': 'Lỗi khi lưu lịch học vào database'},
                    case_name="ERROR",
                    status=500
                )

            # Lấy danh sách lớp đã xếp để trả về
            scheduled_courses = LopTinChi.objects.filter(
                semester_id=semester_id,
                is_deleted=False,
                teacher_id__isnull=False
            ).select_related('teacher', 'teacher__user', 'room', 'subject', 'class_st')

            # Tính toán thống kê
            statistics = self._calculate_statistics(best_chromosome)

            # Chuẩn bị response data
            response_data = {
                'success': True,
                'message': f'Xếp lịch thành công cho {len(best_chromosome.genes)} lớp',
                'scheduled_courses': CourseScheduleSerializer(scheduled_courses, many=True).data,
                'statistics': statistics,
                'violations': best_chromosome.violations,
                'fitness_score': best_chromosome.fitness,
                'generations_run': result['generations'],
                'hard_assignments_count': len(hard_assignments)
            }

            return ResponseFormat.response(data=response_data, case_name="SUCCESS")

        except Exception as e:
            logger.error(f"Lỗi khi xếp lịch: {str(e)}", exc_info=True)
            return ResponseFormat.response(
                data={'message': f'Lỗi: {str(e)}'},
                case_name="ERROR",
                status=500
            )

    def _parse_excel_hard_assignments(self, excel_file):
        """
        Parse file Excel để lấy danh sách hard assignments

        Format Excel:
        | course_id | teacher_id |
        |-----------|------------|
        | 1         | 5          | 
        | 2         | 3          |
        """
        hard_assignments = []

        try:
            # Đọc file Excel
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # Bỏ qua header row
            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            for idx, row in enumerate(rows, start=2):
                # Skip empty rows
                if not any(row):
                    continue

                try:
                    # Parse row data
                    assignment_data = {
                        'course_id': int(row[0]) if row[0] else None,
                        'teacher_id': int(row[1]) if row[1] else None,
                    }

                    # Validate data
                    serializer = HardAssignmentSerializer(data=assignment_data)
                    if not serializer.is_valid():
                        logger.warning(f"Dòng {idx} không hợp lệ: {serializer.errors}")
                        continue

                    # Tạo HardAssignment object
                    assignment = HardAssignment(
                        id=idx,
                        course_id=assignment_data['course_id'],
                        teacher_id=assignment_data['teacher_id'],
                    )

                    hard_assignments.append(assignment)
                    logger.debug(f"Đọc thành công dòng {idx}: Course {assignment.course_id}")

                except (ValueError, TypeError, IndexError) as e:
                    logger.warning(f"Lỗi parse dòng {idx}: {str(e)}")
                    continue

        

        except Exception as e:
            logger.error(f"Lỗi khi đọc file Excel: {str(e)}")
            raise ValueError(f"Không thể đọc file Excel: {str(e)}")

        return hard_assignments


    def _parse_excel_holidays(self, excel_file):
        holidays = set()
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            # ƯU TIÊN CHUẨN VIỆT NAM: NGÀY/THÁNG/NĂM
            date_formats = [
                "%d/%m/%Y",   # 02/09/2025 → 2 tháng 9
                "%d.%m.%Y",
                "%d-%m-%Y",
                "%Y-%m-%d",   # 2025-09-02
            ]

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                raw_value = row[0]
                if raw_value is None or raw_value == "":
                    continue

                # BƯỚC QUAN TRỌNG: LUÔN ÉP VỀ CHUỖI, KHÔNG TIN openpyxl
                s = str(raw_value).strip()

                # Nếu là số (serial date), chuyển về chuỗi dd/mm/yyyy
                if isinstance(raw_value, (int, float)):
                    try:
                        excel_date = datetime(1899, 12, 30) + timedelta(days=int(raw_value))
                        s = excel_date.strftime("%d/%m/%Y")
                    except:
                        s = str(raw_value)

                parsed_date = None
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(s, fmt).date()
                        # Kiểm tra tính hợp lý: nếu ngày > 12 → phải là %d/%m/%Y
                        if fmt == "%d/%m/%Y" and 1 <= parsed_date.day <= 31 and 1 <= parsed_date.month <= 12:
                            break
                        if parsed_date.month > 12:
                            continue
                        break
                    except ValueError:
                        continue

                if parsed_date:
                    holidays.add(parsed_date)
                    # print(f"Parsed: {s} → {parsed_date}")  # Debug
                else:
                    logger.warning(f"[Holiday] Không parse được dòng {idx}: '{s}'")

        except Exception as e:
            raise ValueError(f"Lỗi file Excel: {e}")

        return holidays
    

    def _calculate_statistics(self, chromosome):
        """Tính toán thống kê từ chromosome"""
        DAYS = ["Monday", "Tuesday","Wednesday","Thusday","Friday","Saturday"]
        SLOTS = {
            1: "7:00-11:00",
            6: "13:00-17:00",
        }

        # Thống kê theo ngày
        day_count = {}
        for gene in chromosome.genes:
            day = gene['day_idx']
            day_count[day] = day_count.get(day, 0) + 1
        day_distribution = {
            DAYS[day_idx]: count
            for day_idx, count in sorted(day_count.items())
        }

        # Thống kê theo tiết
        slot_count = {}
        for gene in chromosome.genes:
            slot = gene['slot']
            slot_count[slot] = slot_count.get(slot, 0) + 1

        slot_distribution = {
            f"Tiết {slot} ({SLOTS[slot]})": count
            for slot, count in sorted(slot_count.items())
        }

        # Thống kê giảng viên
        teacher_count = {}
        for gene in chromosome.genes:
            t_id = gene['teacher_id']
            teacher_count[t_id] = teacher_count.get(t_id, 0) + 1

        # Tính trung bình lớp/giảng viên
        avg_classes_per_teacher = (
            sum(teacher_count.values()) / len(teacher_count)
            if teacher_count else 0
        )

        return {
            'total_scheduled': len(chromosome.genes),
            'day_distribution': day_distribution,
            'slot_distribution': slot_distribution,
            'teachers_assigned': len(teacher_count),
            'avg_classes_per_teacher': round(avg_classes_per_teacher, 2),
            'max_classes_teacher': max(teacher_count.values()) if teacher_count else 0,
            'min_classes_teacher': min(teacher_count.values()) if teacher_count else 0
        }


class ScheduleStatusView(APIView):
    """
    API kiểm tra trạng thái xếp lịch của học kỳ
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, semester_id):
        """
        Lấy thông tin trạng thái xếp lịch

        Returns:
        - total_courses: Tổng số lớp trong học kỳ
        - scheduled_courses: Số lớp đã xếp lịch
        - unscheduled_courses: Số lớp chưa xếp lịch
        - schedule_list: Danh sách lớp đã xếp
        """
        try:
            # Tổng số lớp
            total_courses = LopTinChi.objects.filter(
                semester_id=semester_id,
                is_deleted=False
            ).count()

            # Số lớp đã xếp
            scheduled_courses_qs = LopTinChi.objects.filter(
                semester_id=semester_id,
                is_deleted=False,
                teacher_id__isnull=False,
                room_id__isnull=False
            ).select_related('teacher', 'teacher__user', 'room', 'subject', 'class_st')

            scheduled_count = scheduled_courses_qs.count()
            unscheduled_count = total_courses - scheduled_count

            # Serialize danh sách lớp đã xếp
            schedule_list = CourseScheduleSerializer(scheduled_courses_qs, many=True).data

            data = {
                'semester_id': semester_id,
                'total_courses': total_courses,
                'scheduled_courses': scheduled_count,
                'unscheduled_courses': unscheduled_count,
                'completion_rate': round((scheduled_count / total_courses * 100), 2) if total_courses > 0 else 0,
                'schedule_list': schedule_list
            }

            return ResponseFormat.response(data=data, case_name="SUCCESS")

        except Exception as e:
            logger.error(f"Lỗi khi lấy trạng thái xếp lịch: {str(e)}")
            return ResponseFormat.response(
                data={'message': f'Lỗi: {str(e)}'},
                case_name="ERROR",
                status=500
            )


class ScheduleResetView(APIView):
    """
    API xóa lịch học đã xếp
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, semester_id):
        """
        Reset lịch học của học kỳ (xóa teacher, room, weekday, start_period)
        """
        from apps.my_built_in.models.buoi_hoc import BuoiHoc
        try:
            with transaction.atomic():
                BuoiHoc.objects.filter(
                course__semester_id=semester_id
                ).delete()[0]

                updated = LopTinChi.objects.filter(
                    semester_id=semester_id,
                    is_deleted=False
                ).update(
                    teacher_id=None,
                    room_id=None,
                    weekday=None,
                    start_period=None,
                    start_date= None,
                    end_date = None
                )

            logger.info(f"Đã reset {updated} lớp học của học kỳ {semester_id}")

            return ResponseFormat.response(
                data={
                    'message': f'Đã reset lịch học thành công',
                    'courses_reset': updated
                },
                case_name="SUCCESS"
            )

        except Exception as e:
            logger.error(f"Lỗi khi reset lịch học: {str(e)}")
            return ResponseFormat.response(
                data={'message': f'Lỗi: {str(e)}'},
                case_name="ERROR",
                status=500
            )

if __name__ == "__main__":
    text ="Good evening, Nam"
    print(f"trong cau nay co so chu 'e' la: {text.count('e')} va co {text.count('o')} chu o")